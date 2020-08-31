import xlsxwriter
import requests
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl

excel_data_df = pd.read_excel('Образец.xlsx', sheet_name='Лист1')
container = excel_data_df['№ контейнера'].tolist()
i = 1
workbook = xlsxwriter.Workbook('hello.xlsx')
worksheet = workbook.add_worksheet()
for item in container:
    response = requests.get('http://www.cma-cgm.com/ebusiness/tracking/search?SearchBy=Container&Reference=' + item)
    soup = BeautifulSoup(response.text, "lxml")
    worksheet.write(0, i, soup.findAll("td", class_="is-header js-openrow")[-1].text)
    i += 1
    print(soup.findAll("td", class_="is-header js-openrow")[-1].text)
workbook.close()

#wb = openpyxl.load_workbook('Образец.xlsx') #Открываем тестовый Excel файл
#wb.create_sheet('Sheet1') #Создаем лист с названием "Sheet1"
#worksheet = wb['Sheet1'] #Делаем его активным
#worksheet['B4']='We are writing to B4' #В указанную ячейку на активном листе пишем все, что в кавычках
#wb.save('testdel.xlsx') #Сохраняем измененный файл

#workbook = xlsxwriter.Workbook('hello.xlsx')
#worksheet = workbook.add_worksheet()
#worksheet.write('K2', soup.findAll("td", class_="is-header js-openrow")[-1].text)
#worksheet.write(4, 3, 'Hello world')
#workbook.close()
#print(item)
#print(response.text) # вывод содержимого страницы